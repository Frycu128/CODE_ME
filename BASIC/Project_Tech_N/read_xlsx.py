from openpyxl import load_workbook


def load_data(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    return sheet


def load_data_to_list(filename):
    data = load_data(filename)
    data_list = []

    for row in data.iter_rows():
        line = []
        for cell in row:
            line.append(cell.value)
        data_list.append(line)

    return data_list
